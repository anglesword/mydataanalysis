import numpy as np
import yfinance as yf
import pandas as pd
import openpyxl
import ta
import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from pandas import DataFrame
from scipy.signal import argrelextrema
'''
pyinstaller --onefile --icon=ico.ico --add-data "RAW_STOCKES_LIST.xls;." value_investment.py
打包为exe文件命令：
1. 首先使用 PyInstaller 生成默认的 .spec 文件：
    pyinstaller your_script.py
   这会生成一个 your_script.spec 文件和一个 dist/ 目录。
2. 编辑 your_script.spec 文件，在 datas 列表中加入你需要打包的额外文件。例如，如果你有一个 data.txt 文件需要一起打包，可以修改 spec 文件：

注意：打包的文件名后必须加入 ;.
'''
'''
1. 获取股票价值投资得分
2. 判断股票K线形态（用于短线操作），以及相关重要指标
3. 计算高股息的累计分红
4. 对比多个基金，在一定时期内的涨跌幅度
'''

# 1. 定义标的列表，可同时包含美股和加股（加股后缀 .TO）
# tickers = ["MSFT", "AAPL", "WCP.TO", "PEY.TO", "BCE.TO"]
# tickers = list_stocks_from_xls("RAW_STOCKS_LIST.xls")
tickers = {}
start_date = "2000-01-01"
end_date = str(datetime.date.today())

# 2. 权重设置
weights = {
    'dividend_safety': 0.3,
    'de_ratio': 0.25,
    'cashflow_growth': 0.2,
    'industry_stability': 0.15,
    'valuation_reason': 0.1
}

# K线看涨形态map
K_line_pattern_up_map = {
    'ticker': '代码',
    'is_hammer': '锤头线',
    'is_bullish_sun': '大阳线',
    'is_bullish_engulfing': '看涨吞没线',
    'is_morning_star': '早晨之星',
    'is_piercing_pattern': '刺透形态',
    'isGap': '向上跳空',
    'is_three_white_soldiers': '三白兵',
    'is_double_bottom': '双底'
}

# K线看跌形态map
K_line_pattern_down_map = {
    'ticker': '代码',
    'is_evening_star': '黄昏之星',
    'is_bearish_engulfing': '看跌吞没',
    'is_hanging_man': '吊顶线',
    'is_shooting_star': '长上影线阴线(射击之星)',
    'is_bearish_marubozu': '大阴线',
    'is_bearish_variant': '两阴夹一阳',
    'is_doji': '十字星',
    'is_gravestone_doji': '墓碑线',
    'is_dark_cloud_cover': '乌云盖顶',
    'isGap': '向下跳空',
    'is_three_black_crows': '三只乌鸦',
    'is_head_and_shoulders_top': '头顶肩'
}

# 行业中值市盈率映射表（默认静态值），电信行业属于Utilities
industry_pe_map = {
    "Consumer Defensive": 18.5,
    "Utilities": 16.2,
    "Financial": 12.8,
    "Technology": 25.3,
    "Healthcare": 19.4,
    "Basic Materials": 14.0,
    "Default": 20
}

# 功能函数：转换字典的值为列表
def dict_values_to_list(values : list[list[str]]) -> list[str]:
    records = []
    for items in values:
        for item in items:
            records.append(item)

    return records

# 功能函数：读取字典key，根据股票代码获取股票的行业
def fetch_industry(item : str) -> str:
    for key in tickers.keys():
        items = tickers[key]
        for val in items:
            if val == item:
                return key

    return "ERROR"


# 读取存储股票代码的Excel文件
# 读取股票列表，获取股票代码
def fetch_stocks_from_xls(filename : str) :
    df = pd.read_excel(filename, sheet_name=0)
    for idx, row in df.iterrows():
        if row["Industry"] not in tickers:
            tickers[row["Industry"]] = [row["Ticker"]]
        else:
            ts = tickers[row["Industry"]]
            ts.append(row["Ticker"])

'''
第一部分：获取股票价值投资得分
读取源文件：RAW_STOCKES_LIST.xls 获取股票代码列表
写入文件：value_investment_scores.xlsx 的sheet1=Value Scores
'''

# 3. 获取指标函数
def fetch_metrics(ticker):
    stock = yf.Ticker(ticker)
    info = stock.info
    # 基本面指标
    dividend_yield = info.get('dividendYield', 0) * 100
    payout_ratio  = info.get('payoutRatio', 1)
    de_ratio      = info.get('debtToEquity', 999)
    # 近5年经营现金流增长
    growth = calculate_cashflow_growth(ticker)

    return {
        'ticker': ticker,
        'dividend_yield': dividend_yield,
        'payout_ratio': payout_ratio,
        'de_ratio': de_ratio,
        'cashflow_growth': growth,
        'industry_stability': None,  # 手动或逻辑赋值
        'pe': info.get('trailingPE', None)
    }

def get_valuation_score(pe, industry_pe_median):
    print(f"pe={pe},行业中值={industry_pe_median}")
    if pe is None or pd.isna(pe) or pe=='Infinity':
        return 1  # 缺失或无效值，最低分
    elif pe < industry_pe_median:
        return 5
    elif pe < industry_pe_median * 1.2:
        return 3
    else:
        return 1

def calculate_cashflow_growth(ticker):
    print(f"查询股票：{ticker}的现金流")
    try:
        stock = yf.Ticker(ticker)
        cf_df = stock.cashflow
        if cf_df.empty or 'Operating Cash Flow' not in cf_df.index:
            print(f"[{ticker}] 没有找到经营现金流数据")
            return 0.0

        cf_values = cf_df.loc['Operating Cash Flow']

        cf_values = cf_values.dropna()

        if len(cf_values) >= 2:
            # Yahoo Finance 返回的是反向的年份顺序（最新在前）
            latest = cf_values.iloc[0]
            earliest = cf_values.iloc[-1]

            if earliest == 0:
                print(f"[{ticker}] 初始现金流为0，无法计算增长率")
                return 0.0

            growth = ((latest - earliest) / abs(earliest)) * 100
            return round(growth, 2)
        else:
            print(f"[{ticker}] 经营现金流数据不足两年")
            return 0.0
    except Exception as e:
        print(f"[{ticker}] 现金流获取失败: {e}")
        return 0.0

# 4. 批量获取数据并评分
def fetch_value_scores(values : list[str]) -> list[dict[str:float]]:
    records = []
    for t in values:
        data = fetch_metrics(t)
        # 评分规则
        data['dividend_safety'] = 5 if data['payout_ratio'] <= 0.6 else (3 if data['payout_ratio'] <= 0.8 else 1)
        data['de_score'] = 5 if data['de_ratio'] <= 0.5 else (3 if data['de_ratio'] <= 1.0 else 1)
        data['cf_score'] = 5 if data['cashflow_growth'] >= 10 else (3 if data['cashflow_growth'] >= 5 else 1)
        # 假设行业稳定性和估值评分手动或进一步逻辑
        '''
        stock = yf.Ticker(ticker)
        industry_name = stock.info.get("sector", "")
        data['industry_stability'] = get_industry_stability_score(industry_name)
        '''
        industry_value = fetch_industry(t)
        data['industry_stability'] = get_industry_stability_score(industry_value)  # 示例值，请按行业实际情况设定
        print(f"industry_value={industry_value}")
        industry_pe_median = fetch_industry_pe_median(t, industry_value)  # PE行业中值
        '''
        data['valuation_score'] = 5 if (data['pe'] and data['pe'] < industry_pe_median) else (
            3 if (data['pe'] and data['pe'] < industry_pe_median * 1.2) else 1)
        '''
        data['valuation_score'] = get_valuation_score(data['pe'], industry_pe_median)
        # 计算总分
        score = (
                data['dividend_safety'] * weights['dividend_safety'] +
                data['de_score'] * weights['de_ratio'] +
                data['cf_score'] * weights['cashflow_growth'] +
                data['industry_stability'] * weights['industry_stability'] +
                data['valuation_score'] * weights['valuation_reason']
        )
        data['score'] = score

        # 添加评级
        if score >= 4.2:
            data['grade'] = 'A'
        elif score >= 3.5:
            data['grade'] = 'B'
        else:
            data['grade'] = 'C'

        records.append(data)
    return records

# 设定一个行业转换器，用于返回行业的稳定状态，默认得分3分。  # 行业稳定性映射表（默认静态值）
def get_industry_stability_score(industry_name):
    mapping = {
        "Utilities": 5,
        "Consumer Defensive": 5,
        "Communication Services": 5,  # 这里视为稳定行业,电信行业
        "Financial": 4,
        "Healthcare": 4,
        "Technology": 3,
        "Consumer Cyclical": 2,
        "Basic Materials": 2,
        "Energy": 2,
        "Industrials": 2,
        "Real Estate": 2,
        "Etf Index Funds": 3,   # ETF 指数基金
    }
    return mapping.get(industry_name, 3)  # 默认3分


# 5. 获取行业pe中值，其中tickers需要从预定义的excel中取出，用于获取函数：fetch_value_scores()中industry_pe_median的值
def fetch_industry_pe_median(ticker, industry):
    print(f"查询股票{ticker}的PE行业中值")
    pe = 0.00
    try:
        stock = yf.Ticker(ticker)
        pe = stock.info.get("trailingPE")
    except:
        print(f"查询股票{ticker}的PE行业中值失败")

    if pe:
        return float(np.median(pe))
    else:
        return industry_pe_map.get(industry, industry_pe_map["Default"])

# 6. 输出结果
def export_value_scores_to_excel(records, records_up, records_down, buy_sell_scores, total_dividend,
                                 filename="value_investment_scores.xlsx"):
    # 创建 DataFrame
    df = pd.DataFrame(records)
    df = df.sort_values(by='score', ascending=False)

    # 添加中文列名及备注说明（备注将在 Excel 中作为列名文字注释）
    column_map = {
        'ticker': '代码',
        'dividend_yield': '股息率(%)',
        'payout_ratio': '分红支付率',
        'de_ratio': '债务/权益比',
        'cashflow_growth': '经营现金流增长率(%)',
        'industry_stability': '行业稳定性评分',
        'pe': '市盈率(PE)',
        'dividend_safety': '股息安全评分',
        'de_score': '财务稳健评分',
        'cf_score': '现金流增长评分',
        'valuation_score': '估值评分',
        'score': '综合评分',
        'grade': '评级'
    }
    df = df.rename(columns=column_map)

    # 列备注
    column_notes = {
        '股息率(%)': '股息越高越好，但需结合分红可持续性综合判断',
        '分红支付率': '越低越安全，分红支付率=分红/净利润，低于60%为优',
        '债务/权益比': '💰 财务稳健性，≤0.5得5分，≤1.0得3分，>1.0得1分',
        '经营现金流增长率(%)': '📈 越高越佳，≥10%=5分，5~10%=3分，<5%=1分',
        '行业稳定性评分': '🏭 公用/消费=5，银行/保险/医疗=4，科技/可选消费=3，周期性行业=1~2',
        '市盈率(PE)': '用于估值对比，越低越便宜，结合行业中值评估',
        '股息安全评分': '✅ 分红支付率≤60%=5分，60~80%=3分，>80%=1分',
        '财务稳健评分': '💰 债务/权益比≤0.5=5分，≤1.0=3分，>1.0=1分',
        '现金流增长评分': '📈 经营现金流增长率 ≥10%=5分，5~10%=3分，<5%=1分',
        '估值评分': '💸 PE相对行业中值：<中值=5分，<1.2x=3分，≥1.2x=1分',
        '综合评分': '🏆 权重=股息30%、财务25%、现金流20%、行业15%、估值10%',
        '评级': '根据总得分划分：A=优质（≥4.2），B=良好（3.2~4.2），C=一般（<3.2）'
    }

    # 导出带批注的 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Value Scores"

    # 写入列标题并添加备注
    headers = list(df.columns)
    ws.append(headers)

    for idx, col in enumerate(headers, 1):
        note = column_notes.get(col)
        if note:
            ws.cell(row=1, column=idx).comment = openpyxl.comments.Comment(note, "GPT")

    # 写入数据
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # 继续写入K线-UP的状态
    ws = wb.create_sheet(title="K Line UP")  # Excel sheet 名最多31字符
    df = pd.DataFrame(records_up)
    df = df.rename(columns=K_line_pattern_up_map)
    headers = list(df.columns)
    ws.append(headers)
    for row_up in dataframe_to_rows(df, index=False, header=False):
        ws.append(row_up)

    # 继续写入K线-DOWN的状态
    ws = wb.create_sheet(title="K Line DOWN")  # Excel sheet 名最多31字符
    df = pd.DataFrame(records_down)
    df = df.rename(columns=K_line_pattern_down_map)
    headers = list(df.columns)
    ws.append(headers)
    for row_down in dataframe_to_rows(df, index=False, header=False):
        ws.append(row_down)

    # 继续写入辅助指标数据
    ws = wb.create_sheet(title="Buy and Sell")
    df = pd.DataFrame(buy_sell_scores)
    scores_map = {
        'ticker': '代码',
        'boll_buy_score': 'Boll线买入得分',
        'boll_sell_score': 'Boll线卖入得分',
        'rsi_buy_score': 'RSI买入得分',
        'rsi_sell_score': 'RSI卖出得分',
        'ma_buy_score': '均线买入得分',
        'ma_sell_score': '均线卖出得分',
        'macd_buy_score': 'MACD买入得分',
        'macd_sell_score': 'MACD卖出得分',
        'volumne_score': '平均20日交易量得分',
        'buy_score': '买入得分',
        'sell_score': '卖出得分'
    }
    df = df.rename(columns=scores_map)
    headers = list(df.columns)
    ws.append(headers)
    for row_down in dataframe_to_rows(df, index=False, header=False):
        ws.append(row_down)

    # 总收益率
    ws = wb.create_sheet(title="Total Dividend")
    df = pd.DataFrame(total_dividend)
    headers = list(df.columns)
    ws.append(headers)
    for row_down in dataframe_to_rows(df, index=False, header=False):
        ws.append(row_down)

    # 保存文件
    wb.save(filename)

    # 输出控制台预览
    '''
    print(df[[
        '代码', '股息率(%)', '分红支付率', '债务/权益比', '经营现金流增长率(%)',
        '行业稳定性评分', '市盈率(PE)', '股息安全评分', '财务稳健评分',
        '现金流增长评分', '估值评分', '综合评分', '评级']])
    '''

# ################### END 第一部分######################


'''
第二部分：判断股票K线形态（用于短线操作），以及相关重要指标
写入一个文件的两个sheet中：value_investment_scores.xlsx
1. 短线看涨指标写入：sheet2=K Pattern Up Values
2. 短线看跌指标写入：sheet3=K Pattern Down Values
'''
# #################以下函数为K线形态判断#############################
# 定义锤头线的条件-》看涨
'''
参数i表示：
    len(df) - 1  表示 DataFrame 的最后一行的索引，即最近一天的交易数据
    len(df) - 2  代表 倒数第二行的索引，即 前一天的数据。
'''
def is_hammer(df : DataFrame, i : int) -> bool:
    print(f"df={df}")
    open_price = df['Open'].iloc[i]
    close_price = df['Close'].iloc[i]
    high_price = df['High'].iloc[i]
    low_price = df['Low'].iloc[i]

    # 计算实体长度
    body_length = abs(close_price - open_price)
    # 计算上下影线长度
    upper_shadow = high_price - max(open_price, close_price)
    lower_shadow = min(open_price, close_price) - low_price

    # 锤头线的条件
    if close_price > open_price:
        if (high_price - low_price) * 0.3 > body_length > upper_shadow and lower_shadow > body_length * 2:
            return True
    return False

# 定义大阳线的条件-》看涨
def is_bullish_sun(df : DataFrame, i : int) -> bool:
    open_price = df['Open'].iloc[i]
    close_price = df['Close'].iloc[i]
    high_price = df['High'].iloc[i]
    low_price = df['Low'].iloc[i]

    # 计算实体长度
    body_length = abs(close_price - open_price)
    # 计算当天的最高价与收盘价之间的差距
    upper_shadow = high_price - close_price

    # 大阳线的条件
    if close_price > open_price and body_length > (high_price - low_price) * 0.5 and upper_shadow < body_length:
        return True
    return False

# 定义看涨吞没形态的条件-》看涨
'''
看涨吞没的条件：
第一根K线：必须是阴线（即收盘价低于开盘价）。

第二根K线：必须是阳线（即收盘价高于开盘价）。

第二根K线的开盘价低于第一根K线的收盘价，且第二根K线的收盘价高于第一根K线的开盘价。

第二根K线完全吞没第一根K线的实体，即第二根K线的实体范围完全覆盖第一根K线的实体范围。
'''
def is_bullish_engulfing(df, i):
    # 获取前一天和当天的K线数据
    open_price1 = df['Open'].iloc[i - 1]
    close_price1 = df['Close'].iloc[i - 1]
    open_price2 = df['Open'].iloc[i]
    close_price2 = df['Close'].iloc[i]

    # 看涨吞没的条件
    if close_price1 < open_price1 and close_price2 > open_price2:  # 第一根K线是阴线，第二根K线是阳线
        if open_price2 < close_price1 and close_price2 > open_price1:  # 第二根K线完全覆盖第一根K线的实体
            return True
    return False

# 判断K线形态是否“早晨之星” -> 看涨
def is_morning_star(df : DataFrame, day=5, extra=False) -> bool:
    if day < 5:
        return False
    # 获取三天的 K 线数据
    open1, close1 = df['Open'].iloc[day - 2], df['Close'].iloc[day - 2]  # 第一天
    open2, close2, high2, low2 = df['Open'].iloc[day - 1], df['Close'].iloc[day - 1], df['High'].iloc[day - 1], \
            df['Low'].iloc[day - 1]  # 第二天
    open3, close3 = df['Open'].iloc[day], df['Close'].iloc[day]  # 第三天

    # 计算短期趋势（前5天是否处于下跌趋势）
    past_5_days = df['Close'].iloc[day - 5:day]  # 过去5天的收盘价
    downtrend = past_5_days.is_monotonic_decreasing  # 判断是否单调递减

    # 早晨之星形态条件
    cond1 = close1 < open1  # 第一日为大阴线
    cond2 = min(open2, close2) > close1  # 第二日低开
    cond3 = abs(close2 - open2) < (close1 - open1) * 0.3  # 第二日实体较小（小K线或十字星）
    cond4 = close3 > open3  # 第三日为大阳线
    cond5 = close3 > (open1 + close1) / 2  # 第三日收盘价高于第一日实体的一半
    cond6 = downtrend  # 额外条件：前5天是下跌趋势

    if not extra:
        return cond1 and cond2 and cond3 and cond4 and cond5
    else:
        return cond1 and cond2 and cond3 and cond4 and cond5 and cond6


# 刺透形态（Piercing Pattern）
'''
刺透形态（Piercing Pattern） 是一种 看涨反转 的 K 线形态，通常出现在 下降趋势的底部，表示市场可能由空头转向多头。
	•	含义: 买盘反攻，短期看涨，可能反弹。
	•	成交量: 第二日放量更可靠。
'''


def is_piercing_pattern(df: DataFrame) -> bool:
    # 计算前一天的K线数据
    df['Prev Close'] = df['Close'].shift(1)
    df['Prev Open'] = df['Open'].shift(1)
    df['Prev Low'] = df['Low'].shift(1)

    # 计算K线实体部分
    df['Prev Body'] = df['Prev Close'] - df['Prev Open']  # 前一天的实体
    df['Curr Body'] = df['Close'] - df['Open']  # 当天的实体

    # 计算前一天实体的中点
    df['Prev Midpoint'] = df['Prev Open'] + df['Prev Body'] / 2  # 实体中点

    # 识别刺透形态（Piercing Pattern）
    df['Piercing Pattern'] = (
            (df['Prev Body'] < 0) &  # 前一天必须是阴线
            (df['Curr Body'] > 0) &  # 当天必须是阳线
            (df['Open'] < df['Prev Low']) &  # 当天低开
            (df['Close'] > df['Prev Midpoint'])  # 收盘价高于前一天实体中点
    )
    # 获取数据
    latest = df.iloc[-1]
    print(f"刺透形态:{latest['Piercing Pattern']}")
    return latest['Piercing Pattern']


# 跳空缺口（Gap Up/Down）
'''
跳空缺口（Gap） 是指 当前K线的最低价（或最高价）与前一K线的最高价（或最低价）之间出现价格断层，即市场未发生交易而直接跳开。
    1️⃣ 向上跳空（Gap Up）
    当前K线的 最低价高于前一K线的最高价。

    市场买盘强劲，可能预示上涨。

    2️⃣ 向下跳空（Gap Down）
    当前K线的 最高价低于前一K线的最低价。

    市场卖盘强劲，可能预示下跌。
    ◦	跳空高开后回落（阴线）：看跌，顶部压力。
	◦	跳空低开后反弹（阳线）：看涨，底部支撑。
	•	成交量: 高成交量确认趋势。
'''
def is_gap(df: DataFrame) -> list[bool]:
    # 计算前一天的K线数据
    df['Prev High'] = df['High'].shift(1)
    df['Prev Low'] = df['Low'].shift(1)

    # 识别 **向上跳空（Gap Up）**
    df['Gap Up'] = df['Low'] > df['Prev High']

    # 识别 **向下跳空（Gap Down）**
    df['Gap Down'] = df['High'] < df['Prev Low']
    # 获取数据
    latest = df.iloc[-1]
    print(f"跳空缺口->向上跳空:{latest['Gap Up']}, 向下跳空->{latest['Gap Down']}")
    result = [latest['Gap Up'], latest['Gap Down']]
    return result

# 三白兵（Three White Soldiers）
'''
    特征: 连续三根阳线，每日收盘高于前日高点，出现在下跌趋势后。
	•	含义: 买盘持续增强，短期上涨。
	•	成交量: 逐步放大更可靠。
'''
def is_three_white_soldiers(df : DataFrame) -> bool:
    # 🔹 计算 K 线形态
    df['Green1'] = df['Close'] > df['Open']  # 第 1 天阳线
    df['Green2'] = df['Close'].shift(1) > df['Open'].shift(1)  # 第 2 天阳线
    df['Green3'] = df['Close'].shift(2) > df['Open'].shift(2)  # 第 3 天阳线

    # 🔹 确保每根 K 线的开盘价在前一天实体部分之内
    df['Open2 in Body1'] = df['Open'].shift(1) > df['Open'].shift(2)  # 第 2 天开盘 > 第 1 天开盘
    df['Open3 in Body2'] = df['Open'] > df['Open'].shift(1)  # 第 3 天开盘 > 第 2 天开盘

    # 🔹 确保收盘价接近最高价（上影线短）
    df['Small Upper Shadow1'] = (df['High'] - df['Close']) < (df['Close'] - df['Open']) * 0.2
    df['Small Upper Shadow2'] = (df['High'].shift(1) - df['Close'].shift(1)) < (
                df['Close'].shift(1) - df['Open'].shift(1)) * 0.2
    df['Small Upper Shadow3'] = (df['High'].shift(2) - df['Close'].shift(2)) < (
                df['Close'].shift(2) - df['Open'].shift(2)) * 0.2

    # 🔹 形态必须出现在下降趋势后
    df['Downtrend'] = df['Close'].shift(5) > df['Close']

    # 🔹 计算三白兵信号
    df['Three White Soldiers'] = (
            df['Green1'] & df['Green2'] & df['Green3'] &  # 三连阳
            df['Open2 in Body1'] & df['Open3 in Body2'] &  # 开盘价在前日实体部分内
            df['Small Upper Shadow1'] & df['Small Upper Shadow2'] & df['Small Upper Shadow3'] &  # 上影线短
            df['Downtrend']  # 之前是下跌趋势
    )
    # 获取数据
    latest = df.iloc[-1]
    print(f"三白兵:{latest['Three White Soldiers']}")
    return latest['Three White Soldiers']

# 双底（Double Bottom） -> 短期看涨
'''
    •	特征: 两个低点（W形），颈线（中间高点）突破后确认，出现在下跌趋势后。
	•	含义: 底部反转，短期看涨，涨幅约底部到颈线的距离。
	•	成交量: 第二个底放量更强。
'''
def is_double_bottom(df : DataFrame) -> bool:
    # 🔹 计算局部极小值（局部低点）
    # 🔹 计算局部极大值（局部高点）
    local_min_idx = argrelextrema(df['Low'].values, np.greater, order=5)[0]  # 获取索引位置
    local_min_dates = df.index[local_min_idx]  # 转换为日期索引

    # 🔹 赋值局部极大值
    df.loc[local_min_dates, 'Local Min'] = df.loc[local_min_dates, 'Low']

    # 🔹 找出双底模式（两个相近的低点）
    bottoms = df.dropna(subset=['Local Min'])

    if len(bottoms) >= 2:
        for i in range(1, len(bottoms)):
            first_bottom = bottoms.iloc[i - 1]
            second_bottom = bottoms.iloc[i]

            # 条件：两个低点相近（误差不超过 5%）
            if abs(first_bottom['Low'] - second_bottom['Low']) / first_bottom['Low'] < 0.05:

                # 计算颈线（两个底部之间的最高点）
                neckline_level = df.loc[first_bottom.name:second_bottom.name, 'High'].max()

                # 确保价格突破颈线
                if df['Close'].iloc[-1] > neckline_level:
                    df['Double Bottom'] = df.index.isin([first_bottom.name, second_bottom.name])
                    print(f"✅  发现双底信号！可能出现上涨趋势。")
                    return True

    return False

# 计算黄昏之星形态 -> 看跌
def is_evening_star(df : DataFrame, day=5) -> bool:
    """ 检测是否为黄昏之星形态，并判断是否出现在上涨趋势顶部 """
    if day < 5:  # 需要至少5天数据
        return False

        # 获取三天的 K 线数据
    open1, close1 = df['Open'].iloc[day - 2], df['Close'].iloc[day - 2]  # 第一天
    open2, close2 = df['Open'].iloc[day - 1], df['Close'].iloc[day - 1]  # 第二天
    open3, close3 = df['Open'].iloc[day], df['Close'].iloc[day]  # 第三天

    # 计算短期趋势（前5天是否处于上涨趋势）
    past_5_days = df['Close'].iloc[day - 5:day]  # 过去5天的收盘价
    uptrend = past_5_days.is_monotonic_increasing  # 判断是否单调递增

    # 黄昏之星形态条件
    cond1 = close1 > open1  # 第一日为大阳线
    cond2 = min(open2, close2) > close1  # 第二日高开
    cond3 = abs(close2 - open2) < (close1 - open1) * 0.3  # 第二日实体较小
    cond4 = close3 < open3  # 第三日为大阴线
    cond5 = close3 < (open1 + close1) / 2  # 第三日收盘价低于第一日实体的一半
    cond6 = uptrend  # 额外条件：前5天是上涨趋势

    return cond1 and cond2 and cond3 and cond4 and cond5 and cond6

# 定义看跌吞没的条件 -> 看跌
'''
第二天是阴线：Close < Open，即第二天收盘价低于开盘价。

第二天的开盘价高于第一天的收盘价，且第二天的收盘价低于第一天的开盘价：即第二天的阴线实体完全覆盖第一天的阳线。
'''
def is_bearish_engulfing(df, i):
    open_price_today = df['Open'].iloc[i]
    close_price_today = df['Close'].iloc[i]
    open_price_yesterday = df['Open'].iloc[i - 1]
    close_price_yesterday = df['Close'].iloc[i - 1]

    # 确保第二天是阴线，第一天是阳线
    if close_price_today < open_price_today and close_price_yesterday > open_price_yesterday:
        # 检查第二天是否完全覆盖第一天的实体
        if open_price_today > close_price_yesterday and close_price_today < open_price_yesterday:
            return True
    return False

# 定义吊颈线的条件 -> 看跌
'''
阴线：Close < Open，即当日是阴线。

下影线：(Low - min(Open, Close)) > 2 * (Open - Close)，下影线大于实体的两倍。

上影线：(High - max(Open, Close)) < (Open - Close)，上影线小于实体的大小。
'''
def is_hanging_man(df, i):
    open_price = df['Open'].iloc[i]
    close_price = df['Close'].iloc[i]
    high_price = df['High'].iloc[i]
    low_price = df['Low'].iloc[i]

    # 计算实体长度
    body_length = abs(close_price - open_price)
    # 计算上下影线长度
    upper_shadow = high_price - max(open_price, close_price)
    lower_shadow = min(open_price, close_price) - low_price

    # 吊颈线的条件
    if close_price < open_price:  # 阴线
        if lower_shadow > body_length * 2 and upper_shadow < body_length:
            return True
    return False

# 长上影线阴线（Shooting Star） 射击之星
'''
长上影线阴线（Shooting Star，射击之星）**是一种常见的 K 线形态，通常出现在上升趋势的顶部，可能预示着市场即将反转向下
'''
def is_shooting_star(df : DataFrame) -> bool:
    # 计算上影线、下影线、实体大小
    df['Upper Shadow'] = df['High'] - df[['Open', 'Close']].max(axis=1)
    df['Lower Shadow'] = df[['Open', 'Close']].min(axis=1) - df['Low']
    df['Body'] = abs(df['Close'] - df['Open'])
    # 识别 Shooting Star 形态
    df['Shooting Star'] = (
            (df['Upper Shadow'] >= 2 * df['Body']) &  # 上影线至少是实体的2倍
            (df['Lower Shadow'] < df['Body']) &  # 下影线较短
            (df['Close'] < df['Open'])  # 收盘价低于开盘价（阴线）
    )
    # 获取数据
    latest = df.iloc[-1]
    print(f"长上影线阴线（Shooting Star） 射击之星:{latest['Shooting Star']}")
    return latest['Shooting Star']

# 大阴线（Bearish Marubozu）
'''
大阴线（Bearish Marubozu） 是技术分析中的一种强烈看跌的 K 线形态，表示卖方完全占据市场主导地位，通常出现在下跌趋势的开始或延续阶段。
'''
def is_bearish_marubozu(df : DataFrame) -> bool:
    # 计算实体部分、上影线和下影线
    df['Body'] = abs(df['Close'] - df['Open'])  # 实体部分
    df['Upper Shadow'] = df['High'] - df[['Open', 'Close']].max(axis=1)  # 上影线
    df['Lower Shadow'] = df[['Open', 'Close']].min(axis=1) - df['Low']  # 下影线

    # 识别 Bearish Marubozu 形态
    df['Bearish Marubozu'] = (
            (df['Close'] < df['Open']) &  # 阴线（收盘价 < 开盘价）
            (df['Upper Shadow'] <= df['Body'] * 0.1) &  # 上影线几乎没有（小于实体的10%）
            (df['Lower Shadow'] <= df['Body'] * 0.1)  # 下影线几乎没有（小于实体的10%）
    )
    # 获取数据
    latest = df.iloc[-1]
    print(f"大阴线:{latest['Bearish Marubozu']}")
    return latest['Bearish Marubozu']

# 两阴夹一阳（Bearish Three-Line Strike Variant）
'''
反弹失败，空头主导，短期看跌
'''
def is_bearish_variant(df : DataFrame) -> bool:
    # 计算每根K线的类型（阴线或阳线）
    df['Candle Type'] = df.apply(lambda row: 'Bullish' if row['Close'] > row['Open'] else 'Bearish', axis=1)

    # 识别 "两阴夹一阳" 形态
    df['Bearish Three-Line Strike Variant'] = (
            (df['Candle Type'].shift(0) == 'Bearish') &  # 第一根阴线
            (df['Candle Type'].shift(1) == 'Bullish') &  # 第二根阳线
            (df['Candle Type'].shift(2) == 'Bearish') &  # 第三根阴线
            (df['Close'].shift(0) < df['Close'].shift(2))  # 第三根阴线收盘价低于第一根阴线收盘价
    )
    # 获取数据
    latest = df.iloc[-1]
    print(f"两阴夹一阳:{latest['Bearish Three-Line Strike Variant']}")
    return latest['Bearish Three-Line Strike Variant']

# 十字星（Doji）
'''
十字星（Doji） 是一种重要的 K线形态，表示市场买卖力量均衡，可能预示着趋势反转或停滞。
在技术分析中，Doji 形态通常出现在趋势的顶部或底部，反映市场的不确定性。
多空力量平衡，可能是反转或延续信号，需结合前趋势判断：
        上涨后：顶部反转（看跌）。
	◦	下跌后：底部反转（看涨）。
	•	成交量: 高成交量增强信号。
'''
def is_doji(df : DataFrame) -> bool:
    # 计算实体部分（开盘价 - 收盘价的绝对值）
    df['Body'] = abs(df['Close'] - df['Open'])

    # 计算波动范围（最高价 - 最低价）
    df['Range'] = df['High'] - df['Low']

    # 识别 Doji 形态
    df['Doji'] = (df['Body'] <= df['Range'] * 0.1)  # 实体部分小于总波动的 10%

    # 获取数据
    latest = df.iloc[-1]
    print(f"十字星:{latest['Doji']}")
    return latest['Doji']

# 墓碑线（Gravestone Doji）
'''
墓碑线（Gravestone Doji） 是一种 看跌反转K线形态，常出现在 上升趋势的顶部，表示市场尝试上涨但最终被空方压制，收盘价接近开盘价的低点，预示可能的下跌。
	•	含义: 多头冲高后被空头打压至低位，强烈看跌信号。
	•	成交量: 高成交量强化顶部信号。
'''
def is_gravestone_doji(df : DataFrame) -> bool:
    # 计算K线实体部分（开盘价 - 收盘价的绝对值）
    df['Body'] = abs(df['Close'] - df['Open'])

    # 计算影线部分
    df['Upper Shadow'] = df['High'] - df[['Open', 'Close']].max(axis=1)  # 上影线 = 最高价 - 最大(开盘价, 收盘价)
    df['Lower Shadow'] = df[['Open', 'Close']].min(axis=1) - df['Low']  # 下影线 = 最小(开盘价, 收盘价) - 最低价

    # 识别墓碑线（Gravestone Doji）
    df['Gravestone Doji'] = (
            (df['Body'] <= df['High'] * 0.02) &  # 实体部分接近 0
            (df['Upper Shadow'] >= df['Body'] * 5) &  # 长上影线
            (df['Lower Shadow'] <= df['Body'] * 0.5)  # 无下影线或极短
    )
    # 获取数据
    latest = df.iloc[-1]
    print(f"墓碑线:{latest['Gravestone Doji']}")
    return latest['Gravestone Doji']

# 乌云盖顶（Dark Cloud Cover）
'''
乌云盖顶（Dark Cloud Cover） 是 看跌反转 的 K 线形态，通常出现在 上升趋势的顶部，表明市场买盘减弱，空头开始主导市场。
    •	含义: 空头反攻，短期看跌，可能引发回调。
	•	成交量: 第二日放量更强。
'''
def is_dark_cloud_cover(df : DataFrame) -> bool:
    # 计算前一天的K线数据
    df['Prev Close'] = df['Close'].shift(1)
    df['Prev High'] = df['High'].shift(1)
    df['Prev Open'] = df['Open'].shift(1)

    # 计算K线实体部分
    df['Prev Body'] = df['Prev Close'] - df['Prev Open']  # 前一天的实体
    df['Curr Body'] = df['Close'] - df['Open']  # 当天的实体

    # 计算前一天实体中点
    df['Prev Midpoint'] = df['Prev Open'] + df['Prev Body'] / 2  # 实体中点

    # **标准乌云盖顶**
    df['Dark Cloud Cover (Strict)'] = (
            (df['Open'] > df['Prev High']) &  # **严格要求：当天高开**
            (df['Close'] < df['Prev Midpoint']) &  # 收盘跌破前一天实体的一半
            (df['Curr Body'] < 0)  # 当天是阴线
    )

    # **宽松版乌云盖顶**
    df['Dark Cloud Cover (Loose)'] = (
            (df['Open'] > df['Prev Close']) &  # **宽松版：当天开盘价高于前一天收盘价**
            (df['Close'] < df['Prev Midpoint']) &  # 收盘跌破前一天实体的一半
            (df['Curr Body'] < 0)  # 当天是阴线
    )

    # 获取数据
    latest = df.iloc[-1]
    print(f"乌云盖顶:{latest['Dark Cloud Cover (Strict)']}")
    return latest['Dark Cloud Cover (Strict)']

# 三只乌鸦（Three Black Crows）
'''
	•	特征: 连续三根阴线，每日收盘低于前日低点，出现在上升趋势后。
	•	含义: 强烈卖压，短期持续下跌。
	•	成交量: 逐步放大更强。
成交量放大（可选）：
如果 成交量增加，则该形态更具参考价值。
'''
def is_three_black_crows(df : DataFrame) -> bool:
    # 计算 K 线实体
    df['Body1'] = df['Close'].shift(1) - df['Open'].shift(1)  # 前一天实体
    df['Body2'] = df['Close'].shift(2) - df['Open'].shift(2)  # 前二天实体
    df['Body3'] = df['Close'].shift(3) - df['Open'].shift(3)  # 前三天实体

    # 计算三天的开盘和收盘价
    df['Open1'] = df['Open'].shift(1)
    df['Open2'] = df['Open'].shift(2)
    df['Open3'] = df['Open'].shift(3)

    df['Close1'] = df['Close'].shift(1)
    df['Close2'] = df['Close'].shift(2)
    df['Close3'] = df['Close'].shift(3)

    # 计算前 5 天的趋势（判断是否为上涨趋势）
    df['Trend'] = df['Close'].shift(5) < df['Close']

    # 计算移动平均线
    df['SMA5'] = df['Close'].rolling(window=5).mean()  # 5 日均线
    df['SMA10'] = df['Close'].rolling(window=10).mean()  # 10 日均线

    # 识别 **三只乌鸦（Three Black Crows）**
    df['Three Black Crows'] = (
            (df['Body3'] < 0) & (df['Body2'] < 0) & (df['Body1'] < 0) &  # 连续 3 天阴线
            (df['Open2'] < df['Close3']) & (df['Open1'] < df['Close2']) &  # 开盘价在前一天实体内
            (df['Close2'] < df['Close3']) & (df['Close1'] < df['Close2']) &  # 收盘价逐步走低
            (df['Trend']) &  # 前期为上升趋势
            (df['SMA5'] > df['SMA10'])  # 短期均线 > 长期均线（确认上涨趋势）
    )
    # 获取数据
    latest = df.iloc[-1]
    print(f"三只乌鸦:{latest['Three Black Crows']}")
    return latest['Three Black Crows']

# 头肩顶（Head and Shoulders Top）-> 短期看跌
'''
    通常出现在上升趋势的末端
	•	特征: 三段走势：左肩（阳线高点）、头（更高高点）、右肩（低于头的高点），颈线（两低点连线）跌破后确认。
	•	含义: 顶部反转，短期看跌，跌幅约头部到颈线的距离。
	•	成交量: 左肩和头放量，右肩缩量。
'''
def is_head_and_shoulders_top(df : DataFrame) -> bool:
    # 🔹 计算简单移动平均线（SMA）
    df['SMA5'] = df['Close'].rolling(window=5).mean()  # 5日简单移动平均线
    df['SMA20'] = df['Close'].rolling(window=20).mean()  # 20日简单移动平均线

    # 🔹 判断是否处于上升趋势
    # 如果 5日均线高于 20日均线，并且最近的价格也在上升趋势中，则认为是上升趋势
    in_uptrend = df['SMA5'].iloc[-1] > df['SMA20'].iloc[-1]  # 当前 SMA5 大于 SMA20
    if not in_uptrend:
        print("当前未处于上升趋势中...")
        return False
    else:
        # 🔹 计算局部极大值（局部高点）
        local_max_idx = argrelextrema(df['High'].values, np.greater, order=5)[0]  # 获取索引位置
        local_max_dates = df.index[local_max_idx]  # 转换为日期索引

        # 🔹 赋值局部极大值
        df.loc[local_max_dates, 'Local Max'] = df.loc[local_max_dates, 'High']

        # 🔹 识别头肩顶模式（左肩、头部、右肩）
        peaks = df.dropna(subset=['Local Max'])  # 去掉缺失值的行，获取有效的局部极大值
        if len(peaks) >= 3:
            for i in range(1, len(peaks) - 1):
                left_shoulder = peaks.iloc[i - 1]
                head = peaks.iloc[i]
                right_shoulder = peaks.iloc[i + 1]

                # 条件：头部高于左肩和右肩，左右肩高点相近
                if (head['High'] > left_shoulder['High']) and (head['High'] > right_shoulder['High']) and \
                        (abs(left_shoulder['High'] - right_shoulder['High']) / left_shoulder['High'] < 0.05):

                    # 计算颈线（连接左肩和右肩回落点）
                    neckline_level = (left_shoulder['Low'] + right_shoulder['Low']) / 2

                    # 确保价格跌破颈线
                    if df['Close'].iloc[-1] < neckline_level:
                        df['Head and Shoulders'] = df.index.isin([left_shoulder.name, head.name, right_shoulder.name])
                        print(f"⚠️  发现头肩顶信号！可能出现下跌趋势。")
                        return True

    return False
# ################# END K线形态函数#############################

# 获取指定股票数据
def get_stock_data(symbol : str, period="max", interval="1d") -> DataFrame:
    stock = yf.Ticker(symbol)
    df = stock.history(period=period, interval=interval)
    return df

# 返回一只股票的K线看涨形态的指标
def report_k_line_up(symbol : str, df : DataFrame) -> dict[str:list[bool]]:
    line = {"ticker" : symbol, "is_hammer" : is_hammer(df, len(df) - 1), 'is_bullish_sun' : is_bullish_sun(df, len(df) - 1),
            'is_bullish_engulfing' : is_bullish_sun(df, len(df) - 1), 'is_morning_star' : is_morning_star(df),
            'is_piercing_pattern' : is_piercing_pattern(df), 'isGap' : is_gap(df)[0],
            'is_three_white_soldiers' : is_three_white_soldiers(df), 'is_double_bottom' : is_double_bottom(df)}
    return line

# 返回一只股票的K线看跌形态的指标
def report_k_line_down(symbol : str, df : DataFrame) -> dict[str:bool]:
    line = {'ticker': symbol, 'is_evening_star': is_evening_star(df), 'is_bearish_engulfing': is_bearish_engulfing(df, len(df) - 1),
            'is_hanging_man': is_hanging_man(df, len(df) - 1),'is_shooting_star': is_shooting_star(df),
            'is_bearish_marubozu': is_bearish_marubozu(df), 'is_bearish_variant': is_bearish_variant(df),
            'is_doji': is_doji(df), 'is_gravestone_doji': is_gravestone_doji(df), 'is_dark_cloud_cover': is_dark_cloud_cover(df),
            'isGap': is_gap(df)[1], 'is_three_black_crows': is_three_black_crows(df), 'is_head_and_shoulders_top': is_head_and_shoulders_top(df)}
    return line


# 处理所需股票的K线指标，将每只股票的指标放入list[dict[str:bool],最后将所有指标写入Excel
def deal_with_k_line() -> tuple[list[dict[str:bool]], list[dict[str:bool]]]:
    records_up = []
    records_down = []
    for ticker in dict_values_to_list(tickers.values()):
        df = get_stock_data(ticker)
        records_up.append(report_k_line_up(ticker, df))
        records_down.append(report_k_line_down(ticker,df))
    return records_up, records_down

# ####################辅助指标计算#########################


# 计算均线金叉
def detect_golden_cross(df : DataFrame, short_ma : str, long_ma : str) -> bool:
    """检测短期均线上穿长期均线（金叉）"""
    return (df[short_ma] > df[long_ma]) & (df[short_ma].shift(1) <= df[long_ma].shift(1))

# 计算未突破的金叉
def golden_cross_not_confirmed(df : DataFrame, short_ma : str, long_ma : str) -> bool:
    """ 金叉形成后短期均线仍未突破长期均线 """
    cross_signal = detect_golden_cross(df, short_ma, long_ma)
    return cross_signal & (df[short_ma] < df[long_ma])  # 短期均线仍低于长期均线

# 计算均线死叉
def detect_death_cross(df : DataFrame, short_ma : str, long_ma : str) -> bool:
    """检测短期均线上穿长期均线（金叉）"""
    return (df[short_ma] < df[long_ma]) & (df[short_ma].shift(1) >= df[long_ma].shift(1))

# 计算未突破的死叉
def death_cross_not_confirmed(df : DataFrame, short_ma : str, long_ma : str) -> bool:
    """ 金叉形成后短期均线仍未突破长期均线 """
    cross_signal = detect_golden_cross(df, short_ma, long_ma)
    return cross_signal & (df[short_ma] > df[long_ma])  # 短期均线仍低于长期均线

# BB_lower 价格支撑和阻力位参数： 权重->0.2
'''
1. 价格在支撑：价格<=下布林带，值为1（买入信号），否则为0
2. 价格在阻力：价格>=上布林带，值为1（卖出信号), 否则为0
'''
def support_score(symbol : str, df : DataFrame) -> tuple[int, int]:
    sp_up_score = 0
    sp_down_score = 0
    # 计算布林带
    df['MB'] = df['Close'].rolling(window=20).mean()  # 中轨（20日均线）
    df['STD'] = df['Close'].rolling(window=20).std()  # 标准差
    df['UB'] = df['MB'] + (2 * df['STD'])  # 上轨
    df['LB'] = df['MB'] - (2 * df['STD'])  # 下轨
    # 获取最近一天的数据
    latest_data = df.iloc[-1]
    close_price = float.__round__(latest_data['Close'], 2)
    ub = float.__round__(latest_data['UB'], 2)
    lb = float.__round__(latest_data['LB'], 2)
    # 价格在支撑
    if close_price <= lb:
        sp_up_score = 1
    # 价格在阻力
    if close_price >= ub:
        sp_down_score = 1
    return sp_up_score, sp_down_score

# RSI_buy RSI超买和超卖： 权重->0.2
'''
1. RSI<30, 值为1（买入信号），否则为0
2. RSI>70，值为1（卖出信号），否则为0
'''
def rsi_score(symbol : str, df : DataFrame) -> tuple[int, int]:
    rsi_buy_score = 0
    rsi_sell_score = 0
    # 当下以14日的RSI为计算单位
    df['rsi'] = ta.momentum.RSIIndicator(df['Close'], window=14).rsi()
    latest = df.iloc[-1]
    # RSI<30, 值为1（买入信号），否则为0
    if latest['rsi'] < 30:
        rsi_buy_score = 1
    # RSI>70，值为1（卖出信号），否则为0
    if latest['rsi'] > 70:
        rsi_sell_score = 1

    return rsi_buy_score, rsi_sell_score

# MA_cross 看涨（跌）趋势： 权重->0.3
'''
1. 50日均线>200日均线，值为1（买入信号），否则为0
2. 50日均线<200日均线，值为1（卖出信号），否则为0
'''
def cross_score(symbol : str, df : DataFrame) -> tuple[int, int]:
    cross_buy_score = 0
    cross_sell_score = 0
    df['MA50'] = df['Close'].rolling(window=50).mean()
    df['MA200'] = df['Close'].rolling(window=200).mean()
    # 记录金叉信号
    df['Golden_Cross_50_200'] = detect_golden_cross(df, 'MA50', 'MA200')
    # 记录死叉信号
    df['Death_Cross_50_200'] = detect_death_cross(df, 'MA50', 'MA200')
    # 获取最新一天的数据
    latest_data = df.iloc[-1]
    # 50日均线>200日均线，值为1（买入信号），否则为0
    if not latest_data['Golden_Cross_50_200']:
        cross_buy_score = 0
    else:
        cross_buy_score = 1
    # 50日均线<200日均线，值为1（卖出信号），否则为0
    if not latest_data['Death_Cross_50_200']:
        cross_sell_score = 0
    else:
        cross_sell_score = 1
    return cross_buy_score, cross_sell_score

# MACD_buy 正(负)向动能： 权重->0.2
'''
1. MACD线>信号线，值为1（买入信号），否则为0
2. MACD线<信号线，值为1（卖出信号），否则为0
'''
def macd_score(symbol : str, df : DataFrame) -> tuple[int, int]:
    macd_buy_score = 0
    macd_sell_score = 0
    # 计算MACD及信号线
    macd_object = ta.trend.MACD(df['Close'])
    df['MACD'] = macd_object.macd()
    df['MACD_signal'] = macd_object.macd_signal()
    # 获取最新一天的数据
    latest_data = df.iloc[-1]
    # MACD线>信号线，值为1（买入信号），否则为0
    if latest_data['MACD'] > latest_data['MACD_signal']:
        macd_buy_score = 1
    # MACD线<信号线，值为1（卖出信号），否则为0
    if latest_data['MACD'] < latest_data['MACD_signal']:
        macd_sell_score = 1

    return macd_buy_score, macd_sell_score

# V_strength 高交易量：权重->0.1
'''
交易量>过去20天的平均交易量，值为1，否则为0
'''
def volume_score(symbol : str, df : DataFrame) -> int:
    score = 0
    # 过去20天的平均交易量
    df["MA20"] = df["Volume"].rolling(window=20).mean()
    # 获取最新的成交量和均值
    latest = df.iloc[-1]
    if latest['Volume'] > latest['MA20']:
        score = 1
    return score


# ####################END 辅助指标计算#####################

# 辅助指标处理，计算该股票的最终得分，>=0.7为买入或卖出信号
def indicator_buy_and_sell(values : list[str]) -> list[dict[str:float]]:
    final_scores = []
    for ticker in values:
        df = get_stock_data(ticker)
        sp_buy_score, sp_sell_score = support_score(ticker,df)
        rsi_buy_score, rsi_sell_score = rsi_score(ticker, df)
        cross_buy_score, cross_sell_score = cross_score(ticker, df)
        macd_buy_score, macd_sell_score = macd_score(ticker, df)
        vol_score = volume_score(ticker, df)
        buy = 0.2 * sp_buy_score + 0.2 * rsi_buy_score + 0.3 * cross_buy_score + 0.2 * macd_buy_score + 0.1 * vol_score
        sell = 0.2 * sp_sell_score + 0.2 * rsi_sell_score + 0.3 * cross_sell_score + 0.2 * macd_sell_score + 0.1 * vol_score
        ticker_final = {'ticker' : ticker, 'boll_buy_score' : sp_buy_score, 'boll_sell_score' : sp_sell_score,
                        'rsi_buy_score' : rsi_buy_score, 'rsi_sell_score' : rsi_sell_score, 'ma_buy_score' : cross_buy_score,
                        'ma_sell_score' : cross_sell_score, 'macd_buy_score' : macd_buy_score, 'macd_sell_score' : macd_sell_score,
                        'volumne_score' : vol_score, 'buy_score' : buy, 'sell_score' : sell}
        final_scores.append(ticker_final)
    return final_scores
# ################### END 第二部分######################
'''
第三部分：
    计算高股息股票的每股总收益率
    1. 开始时间：start_date = "2000-01-01"
    2. 结束时间：today
'''
def dividend_total_returns() -> list[dict[str:float]]:
    total_returns = []
    for ticker in dict_values_to_list(tickers.values()):
        df = yf.Ticker(ticker)
        # 获取历史收盘价
        hist = df.history(start=start_date, end=end_date)["Close"].dropna()
        if len(hist) < 2:
            print(f"{ticker} 数据不足，跳过")
            continue

        start_price = hist.iloc[0]
        end_price = hist.iloc[-1]

        # 获取历史分红并累计
        dividends = df.dividends[start_date:end_date].sum()

        # 总收益率
        total_return_pct = ((end_price + dividends - start_price) / start_price) * 100

        total_returns.append({
            "ticker" : ticker,
            "开始日期": hist.index[0].date(),
            "结束日期": hist.index[-1].date(),
            "起始价格 ($)": round(start_price, 2),
            "结束价格 ($)": round(end_price, 2),
            "累计分红 ($)": round(dividends, 2),
            "总收益率 (%)": round(total_return_pct, 2)
        })
    return total_returns
# ################### END 第三部分######################
'''
将所有高股息股票的增长趋势，并在excel文件中绘制图形
'''
# 定义函数,获取指定股票在特定日期内的涨跌幅度数据
def fetch_fluctuation_data() -> DataFrame:
    # 下载数据（过去30天）
    data = yf.download(dict_values_to_list(tickers.values()), start=start_date, end=end_date, group_by='ticker', auto_adjust=True)
    # 提取“收盘价”数据
    close_prices = pd.DataFrame({ticker: data[ticker]["Close"] for ticker in dict_values_to_list(tickers.values())})

    # 去除缺失数据
    close_prices.dropna(inplace=True)

    if close_prices.empty:
        print("close_prices 为空，无法计算涨跌幅。")
        return pd.DataFrame()
    # 计算涨跌幅度（相对第一天百分比变化）
    returns = (close_prices - close_prices.iloc[0]) / close_prices.iloc[0] * 100

    # 计算总涨跌幅（最后一天相对于第一天）
    final_returns = (close_prices.iloc[-1] - close_prices.iloc[0]) / close_prices.iloc[0] * 100

    # 排序并打印
    sorted_returns = final_returns.sort_values(ascending=False)
    print("各基金总涨跌幅（%）：\n", sorted_returns)
    print("returns（%）：\n", returns)
    print("final_returns（%）：\n", final_returns)

    return returns
# 写入excel并绘图
def export_trend_to_excel(df : DataFrame, xlsname="value_investment_scores.xlsx") :
    from openpyxl.chart.axis import DateAxis
    sheet_name = '涨跌趋势图'
    chart_position = 'H5'
    # 日期列名
    df = df.copy()
    df.insert(0, "日期", df.index)

    # 尝试打开现有文件，否则新建
    try:
        wb = load_workbook(xlsname)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # 写入数据
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # 创建图表对象


    # 保存
    wb.save(xlsname)

def main() -> None:
    # tickers = list_stocks_from_xls("RAW_STOCKES_LIST.xls")
    fetch_stocks_from_xls("RAW_STOCKES_LIST.xls")
    # 获取估值分数表
    records = fetch_value_scores(dict_values_to_list(tickers.values()))
    # 获取K线短线状态表
    records_up, records_down = deal_with_k_line()
    # 辅助指标处理
    buy_and_sell_scores = indicator_buy_and_sell(dict_values_to_list(tickers.values()))
    # 计算高股息股票的总收益率
    total_dividend = dividend_total_returns()
    # 输出结果到Excel
    export_value_scores_to_excel(records, records_up, records_down, buy_and_sell_scores, total_dividend)
    # 输出图形到Excel
    export_trend_to_excel(fetch_fluctuation_data())
    print(f"程序执行完成！！！")

if __name__ == "__main__":
    main()

